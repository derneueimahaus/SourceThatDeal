"""
File manager for PE Deal Sourcing Automator.
CRUD for local templates (HTML in /templates) with folder support.
CRUD for contact lists (Excel .xlsx in /data).
CRUD for campaigns (JSON in /data/campaigns).
"""

import json
from pathlib import Path

import openpyxl

TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
DATA_DIR = Path(__file__).resolve().parent / "data"
ROOT_FOLDER = "General"  # Templates in templates/*.html live in this logical folder


def _ensure_templates_dir() -> Path:
    """Ensure the templates directory exists; return its path."""
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    return TEMPLATES_DIR


def _normalize_name(name: str) -> str:
    """Ensure name has .html extension for file lookup/write."""
    name = name.strip()
    if not name:
        return ""
    if name.lower().endswith(".html"):
        return name
    return f"{name}.html"


def _display_name(filename: str) -> str:
    """Return display name (without .html) for listing."""
    if filename.lower().endswith(".html"):
        return filename[:-5]
    return filename


def _template_id_to_path(template_id: str) -> Path:
    """Convert template_id (e.g. 'General/My Template' or 'Outreach Q1/Follow-up') to file path."""
    _ensure_templates_dir()
    if "/" in template_id:
        folder, name = template_id.split("/", 1)
        folder, name = folder.strip(), name.strip()
        fname = _normalize_name(name)
        if not fname:
            raise ValueError("Invalid template_id: empty name")
        if folder == ROOT_FOLDER:
            return TEMPLATES_DIR / fname
        return TEMPLATES_DIR / folder / fname
    # Legacy: no slash => treat as General/name
    fname = _normalize_name(template_id)
    if not fname:
        raise ValueError("Invalid template_id: empty name")
    return TEMPLATES_DIR / fname


def _path_to_template_id(folder: str, filename: str) -> str:
    """Build template_id from folder and filename (with .html)."""
    name = _display_name(filename)
    if folder == ROOT_FOLDER or folder == "":
        return f"{ROOT_FOLDER}/{name}"
    return f"{folder}/{name}"


def list_folders() -> list[str]:
    """
    Return folder names: ROOT_FOLDER if any .html in templates/, plus each subdir that contains .html.
    Sorted alphabetically (General first by convention).
    """
    _ensure_templates_dir()
    folders = []
    # Root-level .html files => General
    for p in TEMPLATES_DIR.iterdir():
        if p.is_file() and p.suffix.lower() == ".html":
            folders.append(ROOT_FOLDER)
            break
    # All non-hidden subdirs (including empty ones, so newly created folders appear)
    for p in TEMPLATES_DIR.iterdir():
        if p.is_dir() and not p.name.startswith("."):
            folders.append(p.name)
    return sorted(set(folders), key=lambda f: (0 if f == ROOT_FOLDER else 1, f))


def list_templates(folder: str | None = None) -> list[str]:
    """
    If folder is given: return display names of .html files in that folder (ROOT_FOLDER = root of templates/).
    If folder is None: return all template_ids (Folder/Name) for backward compatibility, flattened.
    Sorted alphabetically within folder.
    """
    _ensure_templates_dir()
    if folder is not None:
        names = []
        if folder == ROOT_FOLDER or folder == "":
            for p in TEMPLATES_DIR.iterdir():
                if p.is_file() and p.suffix.lower() == ".html":
                    names.append(_display_name(p.name))
        else:
            sub = TEMPLATES_DIR / folder
            if sub.is_dir():
                for p in sub.iterdir():
                    if p.is_file() and p.suffix.lower() == ".html":
                        names.append(_display_name(p.name))
        return sorted(names)
    # Flattened: all template_ids (Folder/Name)
    result = []
    for f in list_folders():
        for name in list_templates(f):
            result.append(f"{f}/{name}")
    return sorted(result)


def read_template(template_id: str) -> str:
    """
    Read template content. template_id is "Folder/Name" or legacy "Name" (treated as General/Name).
    Return "" on missing file.
    """
    _ensure_templates_dir()
    try:
        path = _template_id_to_path(template_id)
        return path.read_text(encoding="utf-8")
    except (FileNotFoundError, ValueError):
        return ""
    except Exception:
        raise


def write_template(template_id: str, content: str) -> None:
    """
    Write content to the template. template_id is "Folder/Name". Creates folder if needed.
    """
    _ensure_templates_dir()
    if "/" in template_id:
        folder, name = template_id.split("/", 1)
        folder, name = folder.strip(), name.strip()
        fname = _normalize_name(name)
        if not fname:
            raise ValueError("Template name is empty")
        if folder != ROOT_FOLDER:
            (TEMPLATES_DIR / folder).mkdir(parents=True, exist_ok=True)
            path = TEMPLATES_DIR / folder / fname
        else:
            path = TEMPLATES_DIR / fname
    else:
        fname = _normalize_name(template_id)
        if not fname:
            raise ValueError("Template name is empty")
        path = TEMPLATES_DIR / fname
    path.write_text(content, encoding="utf-8")


def delete_template(template_id: str) -> None:
    """Delete the template file. template_id is "Folder/Name" or "Name"."""
    _ensure_templates_dir()
    try:
        path = _template_id_to_path(template_id)
        path.unlink(missing_ok=True)
    except ValueError:
        pass


def create_folder(folder_name: str) -> None:
    """Create a subfolder under templates/. Fails if name is empty or ROOT_FOLDER."""
    folder_name = folder_name.strip()
    if not folder_name:
        raise ValueError("Folder name is empty")
    if folder_name == ROOT_FOLDER:
        raise ValueError(f"Cannot create folder named '{ROOT_FOLDER}'")
    _ensure_templates_dir()
    (TEMPLATES_DIR / folder_name).mkdir(parents=True, exist_ok=True)


def delete_folder(folder_name: str) -> None:
    """Delete a folder. Only allowed if empty. Cannot delete ROOT_FOLDER."""
    folder_name = folder_name.strip()
    if not folder_name:
        raise ValueError("Folder name is empty")
    if folder_name == ROOT_FOLDER:
        raise ValueError(f"Cannot delete folder '{ROOT_FOLDER}'")
    path = TEMPLATES_DIR / folder_name
    if not path.is_dir():
        return
    if any(path.iterdir()):
        raise ValueError("Folder is not empty")
    path.rmdir()


def move_template(template_id: str, target_folder: str) -> None:
    """Move a template to another folder. target_folder must exist or be ROOT_FOLDER."""
    target_folder = target_folder.strip()
    content = read_template(template_id)
    if "/" in template_id:
        _, name = template_id.split("/", 1)
        name = name.strip()
    else:
        name = template_id.strip()
    new_id = f"{ROOT_FOLDER}/{name}" if target_folder == ROOT_FOLDER or target_folder == "" else f"{target_folder}/{name}"
    write_template(new_id, content)
    delete_template(template_id)


# ---------------------------------------------------------------------------
# Contact list operations (Excel .xlsx in /data)
# ---------------------------------------------------------------------------

def _ensure_data_dir() -> Path:
    """Ensure the data directory exists; return its path."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    return DATA_DIR


def list_contact_lists() -> list[dict]:
    """
    Return list of contact list metadata sorted by name.
    Each dict: {"name": "Main Insight", "filename": "Main Insight.xlsx", "row_count": 123}
    """
    _ensure_data_dir()
    result = []
    for p in sorted(DATA_DIR.iterdir(), key=lambda x: x.name.lower()):
        if p.is_file() and p.suffix.lower() == ".xlsx":
            try:
                wb = openpyxl.load_workbook(p, read_only=True)
                ws = wb.active
                row_count = max(ws.max_row - 1, 0)  # subtract header row
                wb.close()
            except Exception:
                row_count = 0
            result.append({
                "name": p.stem,
                "filename": p.name,
                "row_count": row_count,
            })
    return result


def read_contact_list(filename: str) -> tuple[list[str], list[dict]]:
    """
    Read an Excel contact list. Returns (columns, rows).
    columns: list of header strings from row 1.
    rows: list of dicts keyed by column name.
    """
    _ensure_data_dir()
    path = DATA_DIR / filename
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    columns = []
    rows = []
    for i, cell in enumerate(ws[1]):
        col_name = str(cell.value or f"Column {i + 1}").strip()
        columns.append(col_name)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(columns):
                row_dict[columns[i]] = str(val) if val is not None else ""
        rows.append(row_dict)
    wb.close()
    return columns, rows


def write_contact_list(filename: str, columns: list[str], rows: list[dict]) -> None:
    """Write columns + rows back to an Excel file in /data."""
    _ensure_data_dir()
    path = DATA_DIR / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    for row_dict in rows:
        ws.append([row_dict.get(col, "") for col in columns])
    wb.save(path)
    wb.close()


def delete_contact_list(filename: str) -> None:
    """Delete a contact list .xlsx file."""
    _ensure_data_dir()
    path = DATA_DIR / filename
    path.unlink(missing_ok=True)


def import_contact_list(file_bytes: bytes, target_name: str) -> None:
    """Save uploaded file bytes as an .xlsx in /data."""
    _ensure_data_dir()
    if not target_name.lower().endswith(".xlsx"):
        target_name = f"{target_name}.xlsx"
    path = DATA_DIR / target_name
    path.write_bytes(file_bytes)


# ---------------------------------------------------------------------------
# Campaign operations (JSON in /data/campaigns)
# ---------------------------------------------------------------------------

CAMPAIGNS_DIR = DATA_DIR / "campaigns"


def _ensure_campaigns_dir() -> Path:
    """Ensure the campaigns directory exists; return its path."""
    CAMPAIGNS_DIR.mkdir(parents=True, exist_ok=True)
    return CAMPAIGNS_DIR


def _sanitize_campaign_filename(name: str) -> str:
    """Convert campaign name to a safe filename."""
    safe = name.strip().replace(" ", "_").replace("/", "_")
    if not safe:
        raise ValueError("Campaign name is empty")
    if not safe.lower().endswith(".json"):
        safe = f"{safe}.json"
    return safe


def list_campaigns() -> list[dict]:
    """
    Return list of campaign metadata sorted by name.
    Each dict has at least: name, filename, status, template_id, contact_list.
    """
    _ensure_campaigns_dir()
    result = []
    for p in sorted(CAMPAIGNS_DIR.iterdir(), key=lambda x: x.name.lower()):
        if p.is_file() and p.suffix.lower() == ".json":
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                result.append({
                    "name": data.get("name", p.stem),
                    "filename": p.name,
                    "status": data.get("status", "draft"),
                    "template_id": data.get("template_id", ""),
                    "contact_list": data.get("contact_list", ""),
                })
            except Exception:
                result.append({
                    "name": p.stem,
                    "filename": p.name,
                    "status": "draft",
                    "template_id": "",
                    "contact_list": "",
                })
    return result


def read_campaign(filename: str) -> dict:
    """Read a campaign JSON file. Returns {} if not found."""
    _ensure_campaigns_dir()
    path = CAMPAIGNS_DIR / filename
    if not path.is_file():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def write_campaign(filename: str, data: dict) -> None:
    """Write campaign dict as JSON."""
    _ensure_campaigns_dir()
    path = CAMPAIGNS_DIR / filename
    data["filename"] = filename
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


def delete_campaign(filename: str) -> None:
    """Delete a campaign JSON file."""
    _ensure_campaigns_dir()
    path = CAMPAIGNS_DIR / filename
    path.unlink(missing_ok=True)
